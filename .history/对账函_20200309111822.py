import xlrd
import os
from shutil import copyfile
from xlutils.copy import copy
import xlwt
import openpyxl
from win32com import client
import win32api
from comtypes.client import CreateObject
from impala.dbapi import connect
import pandas as df
from impala.util import as_pandas
from impala.dbapi import connect
from datetime import datetime


def get_data(sql):  # impala中获取数据
    conn = connect(host='10.80.22.229', port=21050, database='dw', timeout=20)
    cursor = conn.cursor()
    cursor.execute(sql)
    data = as_pandas(cursor)
    return data


def exceltopdf(path):
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = 0
    pdf_path = path.replace('xlsx', 'pdf')
    wb = excel.Workbooks.Open(path)
    ws = wb.Worksheets[0]
    try:
        wb.SaveAs(pdf_path, FileFormat=57)
    except Exception as e:
        print("Failed to convert")
        print(str(e))
    finally:
        wb.Close()
        excel.Quit()

# 生成excel和PDF对账涵


def to_pdf_excel(start_date, end_date, date, name, data_1, data_2, data_3):
    start_date = start_date  # '20190901'#对账涵开始时间
    start_date = start_date.replace('-', '')
    end_date = end_date  # '20200229'#对账涵结束时间
    date = date  # '2020/3/5'#对账涵签署日期
    date = ' '*59 + date
    start_date_cn = start_date[0:4] + '年' + \
        str(int(start_date[4:6])) + '月' + '1日'
    end_date_cn = end_date[0:4] + '年' + \
        str(int(end_date[4:6])) + '月' + str(int(end_date[6:8])) + '日'
    info = '    从{}至{}，我司向贵司采购的资金往来款项、发票开立情况、货物收发状态如下：'.format(
        start_date_cn, end_date_cn)  # 对账涵周期提示信息
    # print(start_date_cn)
    # print(end_date_cn)
    # print(info)
    name = name  # '港东荣物资有限公司'#供应商名称
    data_1 = data_1  # 8485069.59 #本期支款给贵公司金额
    data_2 = data_2  # 8480616.41 #本期结算金额
    data_3 = data_3  # 6957087.02 #本期已收发票
    yukuan = round(data_1 - data_2, 2)  # 余款
    yupiao = round(data_2 - data_3, 2)  # 余票

    excel_name = name + '_' + \
        start_date[0:6]+'01' + '-' + end_date + '_对账涵.xlsx'
    # print(excel_name)
    full_file_name_path = 'D:\\pdf\\pdf\\'+excel_name
    copyfile('D:\\pdf\\模板.xlsx', full_file_name_path)  # 复制到pdf目录并重新命名

    wb = openpyxl.load_workbook(full_file_name_path)
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    sheet['B4'] = name + ':'
    sheet['B5'] = info
    sheet['E9'] = data_1
    sheet['G9'] = data_2
    sheet['G14'] = data_3
    sheet['B21'] = date
    wb.save(full_file_name_path)
    exceltopdf(full_file_name_path)
    result_info = '{},excel|PDF对账涵生成完毕，\n\t对账涵周期:{}-{},\n\t本期支款给贵公司金额:{},\n\t本期结算金额:{},\n\t本期已收发票:{},\n\t余款:{},\n\t余票:{}'.format(
        name, start_date, end_date, data_1, data_2, data_3, yukuan, yupiao)
    print(result_info)


# 步骤封装
def save_files(sql, date, end_date):
    impala_data = get_data(sql)
    date = date
    # print(impala_data)
    if len(impala_data) == 0:
        print('未查询到数据！')  # 未查询到数据，pass
        pass
    else:
        for i in range(len(impala_data)):
            name = impala_data['vendor_name'][i]
            start_date = impala_data['min_po_date'][i]
            data_1 = impala_data['pay_amt'][i]
            data_2 = impala_data['shiti_amt'][i]
            data_3 = impala_data['fapiao_amt'][i]
            # print(name,data_1,data_2,data_3)
            yukuan = round(data_1 - data_2, 2)  # 余款
            yupiao = round(data_2 - data_3, 2)  # 余票
            print(i+1, '正在处理>>', name)
            to_pdf_excel(start_date, end_date, date, name,
                         data_1, data_2, data_3)  # 生成excel和PDF
            write_logs(name, start_date, end_date, data_1,
                       data_2, data_3, yukuan, yupiao)  # 保存操作日志
            print('-'*200)

# 保存操作日志


def write_logs(name, start_date, end_date, data_1, data_2, data_3, yukuan, yupiao):
    # save_files(sql,date)
    file = open('excel_pdf_log.csv', 'a')
    # file.write('供应商名称'+','+'对账开始日期'+','+'对账结束日期'+','+'本期支款给贵公司金额'+','+'本期结算金额'+','+'本期已收发票'+','+'余款'+','+'余票'+','+'操作时间'+ '\n')
    now = datetime.now()
    file.write(name+','+start_date+','+end_date+','+str(data_1)+','+str(data_2) +
               ','+str(data_3)+','+str(yukuan)+','+str(yupiao)+',' + str(now)+'\n')
    file.close()


######################################################################################################################################################################################
sql = '''
	select 
	vendor_name,min_po_date,
	case when pay_amt is null then 0 else pay_amt end pay_amt,
	case when shiti_amt is null then 0 else shiti_amt end shiti_amt,
	case when fapiao_amt is null then 0 else fapiao_amt end fapiao_amt 
	from app.vendor_verify--app.vendor_verify_20200131--app.vendor_verify_20191231--  --app.vendor_verify_20200131
	where vendor_name in (
'湖南翔鹏物贸有限公司'
	)
	-- limit 1
	'''

date = '2020/3/5'  # 对账函签署日期
end_date = '20200305'  # '20200229'#对账涵结束时间


save_files(sql, date, end_date)
